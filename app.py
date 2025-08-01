import os
from flask import Flask, render_template, request, jsonify, send_file
import pandas as pd
import io
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB limit

# Créer le dossier uploads s'il n'existe pas
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

def detect_doublons(df):
    try:
        # Nettoyage des données
        df['Latitude'] = pd.to_numeric(df['Latitude'], errors='coerce')
        df['Longitude'] = pd.to_numeric(df['Longitude'], errors='coerce')
        df_coords = df.dropna(subset=['Latitude', 'Longitude'])
        
        # Détection des doublons
        groupes = df_coords.groupby(['Latitude', 'Longitude'])['Identifiant'].agg(['count', 'nunique'])
        doublons = groupes[groupes['nunique'] > 1]
        
        # Récupération des lignes concernées
        lignes_doublons = []
        for (lat, lon), _ in doublons.iterrows():
            mask = (df_coords['Latitude'] == lat) & (df_coords['Longitude'] == lon)
            lignes_doublons.extend(df_coords[mask].index.tolist())
        
        return list(set(lignes_doublons))
    
    except Exception as e:
        print(f"Erreur dans detect_doublons: {str(e)}")
        return []

def detect_errors(df):
    try:
        cols = {
            '2G': {
                'freq': "fréquences d'émission",
                'tilt': "Tits mécanques et électriques de chaque antenne",
                'pire': "Puissance isotrope rayonnée équivalente (PIRE) dans chaque secteur",
                'ant': "Nombre d'antennes",
                'azim': "azimut du rayonnement maximum dans chaque secteur"
            },
            '3G': {
                'tilt': "Tits mécanques et électriques de chaque antenne.1",
                'pire': "Puissance isotrope rayonnée équivalente (PIRE) dans chaque secteur.1",
                'ant': "Nombre d'antennes MIMO",
                'azim': "Azimut du rayonnement maximum dans chaque secteur"
            },
            '4G': {
                'tilt': "Tits mécanques et électriques de chaque antenne.2",
                'pire': "Puissance isotrope rayonnée équivalente (PIRE) dans chaque secteur.2",
                'ant': "Nombre d'antennes MIMO.1",
                'azim': "Azimut du rayonnement maximum dans chaque secteur.1"
            }
        }

        def parse_values(value):
            if pd.isna(value) or str(value).strip() in ['', 'nan']:
                return None
            try:
                parts = str(value).replace(',', '.').split('/')
                return [float(x.strip()) for x in parts if x.strip()]
            except:
                return None

        errors_data = []

        for idx, row in df.iterrows():
            excel_row = idx + 3  # ligne Excel (header à la ligne 2)
            try:
                freq_2g = parse_values(row.get(cols['2G']['freq']))
                if not freq_2g:
                    continue

                ref_count = len(freq_2g)
                ref_azim = parse_values(row.get(cols['2G']['azim']))

                for gen in ['2G', '3G', '4G']:
                    for field in ['tilt', 'pire', 'ant']:
                        col_name = cols[gen][field]
                        val = parse_values(row.get(col_name))
                        if val and len(val) != ref_count:
                            errors_data.append({
                                'Ligne': excel_row,
                                'Colonne': col_name,
                                'Valeur': row.get(col_name),
                                'Problème': f"{gen} - {field} : {len(val)} ≠ {ref_count}"
                            })

                    azim_col = cols[gen]['azim']
                    azim = parse_values(row.get(azim_col))
                    if azim and ref_azim and azim != ref_azim:
                        errors_data.append({
                            'Ligne': excel_row,
                            'Colonne': azim_col,
                            'Valeur': row.get(azim_col),
                            'Problème': f"{gen} - azimut : {azim} ≠ {ref_azim}"
                        })

            except Exception as e:
                print(f"Erreur à la ligne {excel_row} : {e}")
                continue

        return {'errors': errors_data} if errors_data else {}
    except Exception as e:
        print(f"Erreur globale dans detect_errors: {str(e)}")
        return {}

def process_file(file_path, action, sheet_name=None):
    try:
        # Charger le fichier Excel
        wb = load_workbook(file_path)
        ws = wb.active
        sheet_name = sheet_name or ws.title
        
        # Lire les données avec pandas
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=1)
        
        # Nettoyer les noms de colonnes
        df.columns = [str(col).strip() for col in df.columns]
        
        # Renommer les colonnes pour la détection de doublons
        df = df.rename(columns={
            df.columns[0]: 'Identifiant',
            df.columns[4]: 'Longitude', 
            df.columns[5]: 'Latitude'
        })

        results = {}
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Écrire les données originales
            df.to_excel(writer, index=False, sheet_name='Données_originales')
            
            # Exécuter la fonction demandée
            if action == 'detect_errors':
                error_cells = detect_errors(df)
                if error_cells.get('errors'):
                    errors_df = pd.DataFrame(error_cells['errors'])
                    errors_df.to_excel(writer, index=False, sheet_name='Erreurs_détectées')
                    results['errors'] = error_cells['errors']
            
            elif action == 'detect_duplicates':
                doublons_indices = detect_doublons(df)
                if doublons_indices:
                    doublons_data = []
                    for idx in doublons_indices:
                        row = df.iloc[idx]
                        doublons_data.append({
                            'Ligne': idx + 4,  # Ajustement pour la numérotation Excel
                            'Identifiant': row['Identifiant'],
                            'Latitude': row['Latitude'],
                            'Longitude': row['Longitude']
                        })
                    
                    if doublons_data:
                        doublons_df = pd.DataFrame(doublons_data)
                        doublons_df.to_excel(writer, index=False, sheet_name='Doublons_détectés')
                        results['doublons'] = doublons_data
            
            else:
                return None, None, "Action non reconnue"
            
            # Sauvegarder le fichier avec mise en forme
            wb = writer.book
            if action == 'detect_errors' and 'errors' in results:
                ws = wb['Erreurs_détectées']
                red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
                for row in range(2, len(results['errors']) + 2):
                    for col in range(1, 5):
                        ws.cell(row=row, column=col).fill = red_fill
            
            if action == 'detect_duplicates' and 'doublons' in results:
                ws = wb['Doublons_détectés']
                green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                for row in range(2, len(results['doublons']) + 2):
                    for col in range(1, 5):
                        ws.cell(row=row, column=col).fill = green_fill
        
        output.seek(0)
        return results, output.getvalue(), None
    
    except Exception as e:
        return None, None, str(e)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/get_sheets', methods=['POST'])
def get_sheets():
    if 'file' not in request.files:
        return jsonify({'error': 'Aucun fichier fourni'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Aucun fichier sélectionné'}), 400
    
    try:
        # Sauvegarder temporairement le fichier
        filename = secure_filename(file.filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)
        
        # Lire les noms des feuilles
        wb = load_workbook(file_path, read_only=True)
        sheet_names = wb.sheetnames
        wb.close()
        
        # Supprimer le fichier temporaire
        os.remove(file_path)
        
        return jsonify({'sheets': sheet_names})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/process', methods=['POST'])
def process():
    if 'file' not in request.files:
        return jsonify({'error': 'Aucun fichier fourni'}), 400
    
    file = request.files['file']
    action = request.form.get('action')
    sheet_name = request.form.get('sheet_name')
    
    if file.filename == '':
        return jsonify({'error': 'Aucun fichier sélectionné'}), 400
    
    if action not in ['detect_errors', 'detect_duplicates']:
        return jsonify({'error': 'Action non valide'}), 400
    
    try:
        # Sauvegarder temporairement le fichier
        filename = secure_filename(file.filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)
        
        # Traiter le fichier
        results, file_data, error = process_file(file_path, action, sheet_name)
        
        # Supprimer le fichier temporaire
        os.remove(file_path)
        
        if error:
            return jsonify({'error': error}), 400
        
        # Si aucun problème détecté
        if not results:
            return jsonify({
                'message': 'Aucun problème détecté dans le fichier',
                'results': [],
                'file': None
            })
        
        # Retourner les résultats et le fichier modifié
        return jsonify({
            'results': results,
            'file': file_data.hex()  # Convertir en hex pour le JSON
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/download', methods=['POST'])
def download():
    try:
        file_hex = request.json.get('file')
        if not file_hex:
            return jsonify({'error': 'Aucune donnée de fichier fournie'}), 400
        
        file_data = bytes.fromhex(file_hex)
        return send_file(
            io.BytesIO(file_data),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='resultats_analyse.xlsx'
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)